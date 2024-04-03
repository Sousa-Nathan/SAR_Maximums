import os
import pandas as pd
import tkinter as tk
import openpyxl as op
import PySimpleGUI as sg

from datetime import datetime
from tkinter import filedialog
from openpyxl.styles import (Alignment, Border, Font, NamedStyle, PatternFill,
                             Side)

def reported_excel(filepath):
    """
    Formats the workbook for readability.

    Args:
        filepath (str): String filepath to the raw workbook
    """
    
    workbook = op.load_workbook(filename = filepath)
    
    get_sheet = workbook.sheetnames
    
    header = NamedStyle(name = "header")
    header.font = Font(name = "Arial", sz = 8, bold = True)
    header.border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
    header.alignment = Alignment(horizontal = "center", vertical = "center", wrapText = True)
    header.fill = PatternFill(fill_type = "solid", start_color = "00538DD5")
    
    formatted_cells = NamedStyle(name = "formatted_cells")
    formatted_cells.font = Font(name = "Arial", sz = 8)
    formatted_cells.border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"), top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
    formatted_cells.alignment = Alignment(horizontal = "center", vertical = "center", wrapText = True)
    
    for sheet in get_sheet:
        active_sheet = workbook[sheet]
        dimensions = active_sheet.dimensions
        header_row = active_sheet[1]
        
        active_sheet.column_dimensions["A"].width = 16.86
        active_sheet.column_dimensions["B"].width = 13.72
        active_sheet.column_dimensions["C"].width = 13.72
        active_sheet.column_dimensions["D"].width = 20.58
        active_sheet.column_dimensions["E"].width = 13.72
        active_sheet.column_dimensions["F"].width = 13.72
        active_sheet.column_dimensions["G"].width = 13.72
        active_sheet.column_dimensions["H"].width = 13.72
        active_sheet.column_dimensions["I"].width = 13.72
        active_sheet.column_dimensions["J"].width = 13.72
        active_sheet.column_dimensions["K"].width = 13.72
        active_sheet.column_dimensions["L"].width = 13.72
        active_sheet.column_dimensions["M"].width = 13.72
        active_sheet.column_dimensions["N"].width = 13.72
        active_sheet.column_dimensions["O"].width = 13.72
        active_sheet.column_dimensions["P"].width = 13.72
        active_sheet.column_dimensions["Q"].width = 13.72
        
        for row in active_sheet[f"{dimensions}"]:
            for cell in row:
                cell.style = formatted_cells
                cell.style = formatted_cells
        
        for row in active_sheet["L2:M1000"]:
            for cell in row:
                cell.number_format = "0.0"
        
        for row in active_sheet["N2:Q1000"]:
            for cell in row:
                cell.number_format = "0.000"
        
        for cell in header_row:
            cell.style = header
    
    workbook.save(filename = filepath)

def center_window(primary_window, new_window):
    
    
    primary_window_x, primary_window_y = primary_window.current_location()
    primary_window_width, primary_window_height = primary_window.size
    new_win_width, new_win_height = new_window.size
    
    new_x_location = (primary_window_width // 2) - (new_win_width // 2) + primary_window_x
    new_y_location = (primary_window_height // 2) - (new_win_height // 2) + primary_window_y
    
    new_window.move(new_x_location, new_y_location)
    new_window.location = (new_x_location, new_y_location)
    
    return new_x_location, new_y_location

def data_sort(reported_data, over_limit_dir, over_limit_filename):
    workbook = pd.ExcelFile(reported_data)

    sheet_names = workbook.sheet_names

    remove_list = ["Section 1 Summary", "LTE ULCA 5B", "LTE ULCA 7C", "LTE B41 FCC PC2 (Ex)", "LTE ULCA 41C PC3", "LTE ULCA 41C PC2", "LTE ULCA 48C", "FR1 n41 PC2", "FR1 n77 PC2", "FR1 n78 IC PC2", "FR1 n79 (Narrow) PC2", "Wi-Fi 2.4 GHz", "Wi-Fi 5.2 GHz", "Wi-Fi 5.3 GHz", "Wi-Fi 5.5 GHz", "Wi-Fi 5.8 GHz", "U-NII 5", "U-NII 6", "U-NII 7", "U-NII 8", "Bluetooth (2.4 GHz)", "Bluetooth (NB U-NII 1)", "Bluetooth (NB U-NII 3)", "802.15.4", "802.15.4ab", "NFC", "Author"]
    # wlan_list   = ["Wi-Fi 2.4 GHz", "Wi-Fi 5.2 GHz", "Wi-Fi 5.3 GHz", "Wi-Fi 5.5 GHz", "Wi-Fi 5.8 GHz", "U-NII 5", "U-NII 6", "U-NII 7", "U-NII 8"]
    # bt_list     = ["Bluetooth (2.4 GHz)", "Bluetooth (NB U-NII 1)", "Bluetooth (NB U-NII 3)"]
    # thread_list = ["802.15.4", "802.15.4ab"]

    cellular_sar = [name for index, name in enumerate(sheet_names) if name not in remove_list]
    # wlan_sar     = [name for index, name in enumerate(sheet_names) if name in wlan_list]
    # bt_sar       = [name for index, name in enumerate(sheet_names) if name in bt_list]
    # thread_sar   = [name for index, name in enumerate(sheet_names) if name in thread_list]

    cellular_data = [pd.read_excel(reported_data, sheet_name = name) for index, name in enumerate(cellular_sar)]
    # wlan_data     = [pd.read_excel(reported_data, sheet_name = name) for index, name in enumerate(wlan_sar)]
    # bt_data       = [pd.read_excel(reported_data, sheet_name = name) for index, name in enumerate(bt_sar)]
    # thread_data   = [pd.read_excel(reported_data, sheet_name = name) for index, name in enumerate(thread_sar)]

    cellular_test = [cellular_data[index].insert(0, "Technology", cellular_sar[index], True) for index, name in enumerate(cellular_sar)]

    cell_data_trans_0 = [cellular_data[index][cellular_data[index]["Antenna(s)"] == "ANT 0"] for index, name in enumerate(cellular_data)]
    cell_data_trans_1 = [cellular_data[index][cellular_data[index]["Antenna(s)"] == "ANT 1"] for index, name in enumerate(cellular_data)]
    cell_data_trans_2 = [cellular_data[index][cellular_data[index]["Antenna(s)"] == "ANT 2"] for index, name in enumerate(cellular_data)]
    cell_data_trans_5 = [cellular_data[index][cellular_data[index]["Antenna(s)"] == "ANT 5"] for index, name in enumerate(cellular_data)]
    cell_data_trans_6 = [cellular_data[index][cellular_data[index]["Antenna(s)"] == "ANT 6"] for index, name in enumerate(cellular_data)]
    cell_data_trans_7 = [cellular_data[index][cellular_data[index]["Antenna(s)"] == "ANT 7"] for index, name in enumerate(cellular_data)]

    cell_data_max_0 = [cell_data_trans_0[index].loc[cell_data_trans_0[index].groupby(by = "RF Exposure Condition")["1-g Scaled (W/kg)"].idxmax()].sort_index() for index, name in enumerate(cell_data_trans_0)]
    cell_data_max_1 = [cell_data_trans_1[index].loc[cell_data_trans_1[index].groupby(by = "RF Exposure Condition")["1-g Scaled (W/kg)"].idxmax()].sort_index() for index, name in enumerate(cell_data_trans_1)]
    cell_data_max_2 = [cell_data_trans_2[index].loc[cell_data_trans_2[index].groupby(by = "RF Exposure Condition")["1-g Scaled (W/kg)"].idxmax()].sort_index() for index, name in enumerate(cell_data_trans_2)]
    cell_data_max_5 = [cell_data_trans_5[index].loc[cell_data_trans_5[index].groupby(by = "RF Exposure Condition")["1-g Scaled (W/kg)"].idxmax()].sort_index() for index, name in enumerate(cell_data_trans_5)]
    cell_data_max_6 = [cell_data_trans_6[index].loc[cell_data_trans_6[index].groupby(by = "RF Exposure Condition")["1-g Scaled (W/kg)"].idxmax()].sort_index() for index, name in enumerate(cell_data_trans_6)]
    cell_data_max_7 = [cell_data_trans_7[index].loc[cell_data_trans_7[index].groupby(by = "RF Exposure Condition")["1-g Scaled (W/kg)"].idxmax()].sort_index() for index, name in enumerate(cell_data_trans_7)]

    cell_max_0 = pd.concat(cell_data_max_0)
    cell_max_1 = pd.concat(cell_data_max_1)
    cell_max_2 = pd.concat(cell_data_max_2)
    cell_max_5 = pd.concat(cell_data_max_5)
    cell_max_6 = pd.concat(cell_data_max_6)
    cell_max_7 = pd.concat(cell_data_max_7)

    transmitters = ["ANT 0", "ANT 1", "ANT 2", "ANT 5", "ANT 6", "ANT 7"]
    cell_maximums = [cell_max_0.filter(items = ["Technology", "Antenna(s)", "RF Exposure Condition", "Mode(s)", "Power Mode(s)", "Dist. (mm)", "Test Position(s)", "Channel", "Freq. (MHz)", "RB Allocation", "RB Offset", "Max Output Pwr (dBm)", "Meas. (dBm)", "1-g Meas. (W/kg)", "1-g Scaled (W/kg)", "10-g Meas. (W/kg)", "10-g Scaled (W/kg)"]), cell_max_1.filter(items = ["Technology", "Antenna(s)", "RF Exposure Condition", "Mode(s)", "Power Mode(s)", "Dist. (mm)", "Test Position(s)", "Channel", "Freq. (MHz)", "RB Allocation", "RB Offset", "Max Output Pwr (dBm)", "Meas. (dBm)", "1-g Meas. (W/kg)", "1-g Scaled (W/kg)", "10-g Meas. (W/kg)", "10-g Scaled (W/kg)"]), cell_max_2.filter(items = ["Technology", "Antenna(s)", "RF Exposure Condition", "Mode(s)", "Power Mode(s)", "Dist. (mm)", "Test Position(s)", "Channel", "Freq. (MHz)", "RB Allocation", "RB Offset", "Max Output Pwr (dBm)", "Meas. (dBm)", "1-g Meas. (W/kg)", "1-g Scaled (W/kg)", "10-g Meas. (W/kg)", "10-g Scaled (W/kg)"]), cell_max_5.filter(items = ["Technology", "Antenna(s)", "RF Exposure Condition", "Mode(s)", "Power Mode(s)", "Dist. (mm)", "Test Position(s)", "Channel", "Freq. (MHz)", "RB Allocation", "RB Offset", "Max Output Pwr (dBm)", "Meas. (dBm)", "1-g Meas. (W/kg)", "1-g Scaled (W/kg)", "10-g Meas. (W/kg)", "10-g Scaled (W/kg)"]), cell_max_6.filter(items = ["Technology", "Antenna(s)", "RF Exposure Condition", "Mode(s)", "Power Mode(s)", "Dist. (mm)", "Test Position(s)", "Channel", "Freq. (MHz)", "RB Allocation", "RB Offset", "Max Output Pwr (dBm)", "Meas. (dBm)", "1-g Meas. (W/kg)", "1-g Scaled (W/kg)", "10-g Meas. (W/kg)", "10-g Scaled (W/kg)"]), cell_max_7.filter(items = ["Technology", "Antenna(s)", "RF Exposure Condition", "Mode(s)", "Power Mode(s)", "Dist. (mm)", "Test Position(s)", "Channel", "Freq. (MHz)", "RB Allocation", "RB Offset", "Max Output Pwr (dBm)", "Meas. (dBm)", "1-g Meas. (W/kg)", "1-g Scaled (W/kg)", "10-g Meas. (W/kg)", "10-g Scaled (W/kg)"])]

    with pd.ExcelWriter(over_limit_filename) as writer:
        for data in range(len(transmitters)):
            cell_maximums[data].to_excel(writer, sheet_name = transmitters[data], index = False)

def main_window():
    
    sg.theme("Dark")
    
    file_list_column = [
        [
            sg.Text("Output Directory for Cellular Maximums Workbook: "),
            sg.Input(key = "-CELL_MAX_OUT-"),
            sg.FolderBrowse(key = "-CELL_MAX_OUT_BROWSE-", button_color = ("black", "#D3D3D3"))
        ],
    ]
    
    selected_files_column = [
        [
            sg.Text("Reported Results Workbook:"),
            sg.Input(key = "-Workbook_1-"),
            sg.FileBrowse(key = "-WB1_IN_BROWSE-", button_color = ("black", "#D3D3D3"), file_types = (("Excel File", "*.xlsx*"),))
        ],
    ]
    
    initial_layout = [
        [
            sg.Column(file_list_column, element_justification = "right"),
            sg.VerticalSeparator(),
            sg.Column(selected_files_column, element_justification = "left")
        ],
        
        [
            sg.Exit(button_color = ("white", "red")),
            sg.Button("Clear", key = "-CLEAR-", button_color = ("black", "#D3D3D3")),
            sg.Button("Find Maximums", key = "-MAX-", button_color = ("black", "#D3D3D3")),
        ]
    ]
    
    initial_window = sg.Window("Maximums", initial_layout, finalize = True)
    
    while True:
        event, values = initial_window.read()
        # print(f"Event: {event}; Value: {values}")
        if event in (sg.WINDOW_CLOSED, "Exit"):
            break
        
        if event == "-CLEAR-":
            keys_to_clear = ["-CELL_MAX_OUT-", "-Workbook_1-"]
            
            for key in keys_to_clear:
                initial_window[key].update("")
                values[key] = ""
        
        if event == "-MAX-":
            now = datetime.now()
            date = now.strftime("%Y_%m_%d")
            time = now.strftime("%H_%M_%S")
            sar_max_filename = os.path.join(values["-CELL_MAX_OUT-"], f"Max_SAR_{date}_{time}.xlsx")
            
            data_sort(values["-Workbook_1-"], values["-CELL_MAX_OUT-"], sar_max_filename)
            reported_excel(sar_max_filename)
            
            sg.popup("Done.", button_color = ("black", "#D3D3D3"), location = (center_window(initial_window, new_window = initial_window)))

if __name__ == "__main__":
    main_window()